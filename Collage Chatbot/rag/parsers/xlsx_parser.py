import io
import hashlib
from typing import Dict, Any, Optional, List

class XLSXParser:
    """
    XLSX document parser for extracting text, metadata, and structure from Excel spreadsheets.
    Supports basic text extraction, cell-level processing, and metadata extraction.
    """
    
    def __init__(self):
        self.supported_formats = ['xlsx', 'xls']
        self.openpyxl_available = False
        
        try:
            import openpyxl
            self.openpyxl = openpyxl
            self.openpyxl_available = True
        except ImportError:
            print("[XLSXParser] openpyxl not installed. Install with: pip install openpyxl")
    
    def parse_xlsx(self, file_bytes: bytes, filename: str = "") -> Dict[str, Any]:
        """
        Parse XLSX file and extract text, metadata, and structure.
        
        Args:
            file_bytes: Raw bytes of the XLSX file
            filename: Original filename of the XLSX
            
        Returns:
            Dictionary containing extracted content and metadata
        """
        if not self.openpyxl_available:
            return {
                "success": False,
                "error": "openpyxl not installed",
                "filename": filename
            }
        
        try:
            xls_file = io.BytesIO(file_bytes)
            workbook = self.openpyxl.load_workbook(xls_file, data_only=True)
            
            # Extract basic metadata
            metadata = self._extract_metadata(workbook, filename)
            
            # Extract text from all sheets
            sheets_content = []
            full_text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = ""
                rows_data = []
                
                for row_num, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    row_text = " | ".join(row_values)
                    if row_text.strip():
                        sheet_text += row_text + "\n"
                        rows_data.append({
                            "row_number": row_num + 1,
                            "values": row_values,
                            "text": row_text,
                            "cell_count": len(row_values)
                        })
                
                if sheet_text.strip():
                    sheet_content = {
                        "sheet_name": sheet_name,
                        "text": sheet_text.strip(),
                        "rows": rows_data,
                        "total_rows": len(rows_data),
                        "char_count": len(sheet_text)
                    }
                    sheets_content.append(sheet_content)
                    full_text += f"Sheet: {sheet_name}\n{sheet_text}\n\n"
            
            # Clean and structure the extracted text
            clean_text = self._clean_text(full_text)
            
            # Generate content hash
            content_hash = hashlib.sha256(clean_text.encode('utf-8')).hexdigest()
            
            # Extract sections (each sheet is a section)
            sections = self._extract_sections(sheets_content)
            
            return {
                "success": True,
                "filename": filename,
                "format": "xlsx",
                "metadata": metadata,
                "full_text": clean_text,
                "content_hash": content_hash,
                "sheets": sheets_content,
                "total_sheets": len(sheets_content),
                "total_characters": len(clean_text),
                "sections": sections,
                "word_count": len(clean_text.split())
            }
            
        except Exception as e:
            print(f"[XLSXParser] Error parsing XLSX: {e}")
            return {
                "success": False,
                "error": str(e),
                "filename": filename
            }
    
    def _extract_metadata(self, workbook, filename: str) -> Dict[str, Any]:
        """Extract metadata from XLSX workbook"""
        metadata = {
            "filename": filename,
            "title": "",
            "author": "",
            "subject": "",
            "creator": "",
            "created": "",
            "modified": "",
            "sheet_count": len(workbook.sheetnames)
        }
        
        try:
            props = workbook.properties
            if props:
                metadata.update({
                    "title": props.title or "",
                    "author": props.creator or "",
                    "subject": props.subject or "",
                    "creator": props.creator or "",
                    "created": str(props.created) if props.created else "",
                    "modified": str(props.modified) if props.modified else ""
                })
        except Exception as e:
            print(f"[XLSXParser] Error extracting metadata: {e}")
        
        return metadata
    
    def _clean_text(self, text: str) -> str:
        """Clean and normalize extracted text"""
        if not text:
            return ""
        
        # Remove excessive whitespace
        text = ' '.join(text.split())
        
        # Normalize line breaks
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in text:
            text = text.replace('\n\n\n', '\n\n')
        
        return text.strip()
    
    def _extract_sections(self, sheets_content: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Extract sections from sheets. Each sheet is treated as a section.
        """
        sections = []
        
        for sheet in sheets_content:
            sections.append({
                "title": sheet["sheet_name"],
                "content": sheet["text"],
                "sheet_name": sheet["sheet_name"]
            })
        
        return sections
    
    def is_supported_format(self, filename: str) -> bool:
        """Check if the file format is supported"""
        return filename.lower().endswith(('.xlsx', '.xls'))
    
    def validate_xlsx(self, file_bytes: bytes):
        """
        Validate if the file is a valid XLSX.
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not self.openpyxl_available:
            return False, "openpyxl not installed"
        
        try:
            xls_file = io.BytesIO(file_bytes)
            workbook = self.openpyxl.load_workbook(xls_file, data_only=True)
            
            # Try to read the first sheet to validate
            if len(workbook.sheetnames) > 0:
                first_sheet = workbook[workbook.sheetnames[0]]
                _ = list(first_sheet.iter_rows(values_only=True))
                
            return True, ""
            
        except Exception as e:
            return False, f"Invalid XLSX file: {str(e)}"
